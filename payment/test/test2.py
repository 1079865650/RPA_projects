# _*_ coding : utf-8 _*_
# @Time : 2023-03-30 9:14
# @Author : wws
# @File : test2
# @Project : RPA_projects
import datetime
import os
import re

from openpyxl import Workbook

def write_to_excel(aa: list, site):
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    data_path = r'F:\Users\DeskTop\payment\\' + site + '\\' + today + '\\' + 'data'
    if not os.path.exists(data_path):
        os.makedirs(data_path)
    wb = Workbook()
    ws = wb.active  # 获取当前活跃的sheet，默认为第一张sheet
    for index, data01 in enumerate(aa):
        data = str(data01)
        ws['A' + str(index + 1)] = data
        re_space = re.sub(r'[" "]', '', data)
        ws['B' + str(index + 1)] = re_space
    wb.save(os.path.join(data_path, r'data.xlsx'))