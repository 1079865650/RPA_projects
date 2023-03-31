# _*_ coding : utf-8 _*_
# @Time : 2023-03-29 18:16
# @Author : wws
# @File : test
# @Project : RPA_projects


# import openpyxl
# data = openpyxl.load_workbook(r'F:\Users\DeskTop\payment\payment_file\2023-03-29\data.xlsx')
# print(data.get_named_ranges()) # 输出工作页索引范围
# print(data.get_sheet_names()) # 输出所有工作页的名称
# # 取第一张表
# sheetnames = data.get_sheet_names()
# table = data.get_sheet_by_name(sheetnames[0])
# table = data.active
# print(table.title) # 输出表名
# nrows = table.max_row # 获得行数
# ncolumns = table.max_column # 获得行数
# values = ['E','X','C','E','L']
# for value in values:
#     table.cell(nrows+1,1).value = value
#     nrows = nrows + 1
# data.save(r'F:\Users\DeskTop\payment\payment_file\2023-03-29\data.xlsx')

from openpyxl import load_workbook

a = [1,23,4]
a.pop(len(a)-1)
print(a)
b = a.remove(23)
print(b)
















