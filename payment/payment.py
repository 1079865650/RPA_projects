# _*_ coding : utf-8 _*_
# @Time : 2023-02-23 13:59
# @Author : wws
# @File : test
# @Project : python基础
import datetime
import os
import re
import redis
import openpyxl
from settings import *
from openpyxl import Workbook


# download the file name to be downloaded to Excel
# aa(list)：写入excel的文件名 site:站点
# 根据站点写入到excel1的不同位置，最后通过文件名找需要下载的报表
def write_to_excel(aa, site):
    print("=========filename_list", aa)
    today = datetime.datetime.now().strftime("%Y-%m")
    data_path_splice = os.path.join(origin_path, today)
    if not os.path.exists(data_path_splice):  # create a folder name by month  通过月份创建文件夹
        os.makedirs(data_path_splice)
    excel_data = os.path.join(data_path_splice, r'data.xlsx')
    if not os.path.exists(excel_data):  # create a excel  # 再该文件夹下面创建excel
        wb = Workbook()
        wb.save(excel_data)
        wb.close()
    wb = openpyxl.load_workbook(excel_data)  # load excel driver   # 加载excel驱动
    sheet = wb.active
    avail_row = sheet.max_row  # obtain the maximum row available in Excel  # 获取excel最大行可用行
    site_dict = {'EU': [1, 2], 'US': [3, 4], 'JP': [5, 6]}  # write different columns according to different sites  # 根据站点写到不同的excel列
    now_site = site_dict[site]
    for index, data01 in enumerate(aa):
        data = str(data01)
        sheet.cell(row=avail_row + index + 1, column=now_site[0]).value = data  # write raw data  # 写入数据
        re_space = re.sub(r'[" "]', '', data)
        sheet.cell(row=avail_row + index + 1, column=now_site[1]).value = re_space  # write data after removing spaces # 对数据去空格
    wb.save(excel_data)


# read excel,which contains the file name to be downloaded
# 读取excel获取所有文件名去判断是否下载该文件
def read_excel(site):
    site_dict = {'EU': ['A', 'B'], 'US': ['C', 'D'], 'JP': ['E', 'F']}
    site_now = site_dict[site]  # write different columns according to different sites
    excel_path = os.path.join(origin_path, datetime.datetime.now().strftime("%Y-%m") + '\\' + 'data.xlsx')
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    cell_list = []
    for cell in sheet[site_now[1]]:  # 遍历excel列读取文件名
        if cell.value is not None:
            cell_list.append(cell.value)
    return cell_list


# determined whether to download
# filename(str)：文件名, cell_list(list)：需要下载的excel文件名集合, site(str):站点
# 判断改文件名是否在excel里面，excel里面存放所有需要下载的文件名
def whether_to_download(filename, cell_list, site):
    re_space = re.sub(r'[" "]', '', filename)
    if re_space in cell_list:
        file_path = os.path.join(origin_path, datetime.datetime.now().strftime("%Y-%m") + '\\' + site)
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        return file_path
    return False


# rename the downloaded file according to the site
# site1(str):站点, country(str)：国家, us_category(str):us站点的附加分类
# 对下载的附件根据站点重新命名
def rename_file(site, country, us_category=""):
    country_abbreviation = {'Germany': 'DE', 'France': 'FR', 'Italy': 'IT', 'Spain': 'ES', 'United Kingdom': 'UK',
                            'Poland': 'PL', 'Turkey': 'TR', 'Netherlands': 'NL', 'Belgium': 'BE', 'Sweden': 'SE',
                            'Japan': 'JP', 'United States': 'US', 'Mexico': 'MX', 'Canada': 'CA'}
    abbreviation_list = []
    for key in country_abbreviation.keys():
        abbreviation_list.append(country_abbreviation[key])
    prefix = country_abbreviation[country]
    jp_path = os.path.join(origin_path, datetime.datetime.now().strftime("%Y-%m") + '\\' + site)
    g = os.walk(jp_path)
    absolute_path_list = []
    for path, dir_list, file_list in g:  # 遍历文件夹拿到所有附件路径
        for file_name in file_list:
            if file_name[0:2] in abbreviation_list:  # determine whether it has been renamed   #判断是否需要重新命名
                continue
            if us_category != "":
                us_category = us_category + " "
            modify_file_name = prefix + " " + us_category + file_name
            modify_path = os.path.join(path, modify_file_name)  # modified file name # 创建新路径
            absolute_path = os.path.join(path, file_name)  # raw name
            item = [absolute_path, modify_path]
            absolute_path_list.append(item)
    print(absolute_path_list)
    path_str = ''
    for i in absolute_path_list:
        a = i[0]
        b = i[1]
        str1 = b.split("\\")[-1]
        path_str = str1 + ','
        if os.path.exists(b):
            os.remove(b)  # if the file is exists,delete it  如果该文件存在，删除该文件
        os.rename(a, b)  # 重新命名
    return path_str


# determine the prefix for file naming
# 判断us站点的分类条件
def b2b_b2c(value):
    value_split = value.split(" ")
    if 'Unified' in value_split:
        return 'ALL'
    else:
        return 'B2C'


# redis的增删改查
def add(rd_key, rd_value, expired=-1):
    redis_conn = redis.Redis(host=redis_host, port=redis_port, db=redis_db)
    if expired != -1:
        add_result = redis_conn.setex(rd_key, expired, rd_value)
    else:
        add_result = redis_conn.set(rd_key, rd_value)
    redis_conn.close()
    return add_result


def exists(rd_key):
    redis_conn = redis.Redis(host=redis_host, port=redis_port, db=redis_db)
    exist = redis_conn.exists(rd_key)
    redis_conn.close()
    return True if exist > 0 else False


def delete_key(rd_key):
    # redis_conn = connect_redis()
    redis_conn = redis.Redis(host=redis_host, port=redis_port, db=redis_db)
    delete_result = redis_conn.delete(rd_key)
    redis_conn.close()
    return True if delete_result > 0 else False


def add_redis(country):
    date_str = datetime.datetime.now().strftime("%Y-%m")
    add("payment:before:"+date_str+country, 1, expiration_time)


def add_redis_download(country):
    add("payment:downloaded:"+country, 1, expiration_time)


# 删除所有redis' key
def delete_all_keys():
    redis_conn = redis.Redis(host=redis_host, port=redis_port, db=redis_db)
    keys = redis_conn.keys()
    key_list = []
    for key in keys:
        if 'payment' in bytes(key).decode('utf-8') and 'test' not in bytes(key).decode('utf-8'):
            key_list.append(key)
    for i in key_list:
        redis_conn.delete(i)
    redis_conn.close()
    print("=========key_list  delete key values about redis", key_list)


def judge_all_b2c(string_value):
    string_split = string_value.split(" ")
    if 'Unified' in string_split:
        return 'ALL'
    return 'B2C'


# 移动文件到smb
def remove_file_to_smb():
    # file_path = os.path.join(origin_path, datetime.datetime.now().strftime("%Y-%m"))
    file_path = os.path.join(origin_path, datetime.datetime.now().strftime("2023-03"))
    print(file_path)
    g = os.walk(file_path)
    dir_file = []
    for path, dir_list, file_list in g:
        for i in file_list:
            if i == 'data.xlsx':
                continue
            # absolute_path = os.path.join(path, i)
            folder = path.split("\\")[-1]
            item = [folder, path, i]
            dir_file.append(item)
    print("=========dir_file folder and filename", dir_file)
    return dir_file
#
# a = remove_file_to_smb()
# print(a)


def excel_path():
    return [Excel_EU_path, Excel_US_path, Excel_JP_path]


# 查询redis里面建的数量 判断流程执行到哪一步
def query_keys_number():
    redis_conn = redis.Redis(host=redis_host, port=redis_port, db=redis_db)
    keys = redis_conn.keys()
    key_list_before = []
    key_list_downloaded = []
    pre_key = datetime.datetime.now().strftime("%Y-%m")
    country_list = ['Germany', 'France', 'Italy', 'Spain', 'United Kingdom', 'Poland', 'Turkey', 'Netherlands',
                    'Belgium', 'Sweden', 'Japan', 'United States', 'Mexico', 'Canada']
    before_list = []
    downloaded_list = []
    for i in country_list:
        ii = 'payment:before:' + pre_key + i
        aa = 'payment:downloaded:' + i
        before_list.append(ii)
        downloaded_list.append(aa)
    print(before_list)
    for key in keys:
        key_str = bytes(key).decode('utf-8')
        if key_str in before_list:
            key_list_before.append(key_str)
        if key_str in downloaded_list:
            key_list_downloaded.append(key_str)
    return [len(key_list_before), len(key_list_downloaded)]


a = query_keys_number()
print(a)


# is the time after 5 pm on first day of each month?
def judge_time():
    date_time = datetime.datetime.now().strftime("%Y-%m")
    date_split = date_time.split("-")
    compare_time = datetime.datetime(int(date_split[0]), int(date_split[1]), 1, 17, 0, 0)
    now = datetime.datetime.now()
    return now > compare_time



















