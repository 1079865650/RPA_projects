# _*_ coding : utf-8 _*_
# @Time : 2023-03-30 9:21
# @Author : wws
# @File : demo
# @Project : RPA_projects
import datetime
import os
import re

import openpyxl
from settings import *
from openpyxl import Workbook


def write_to_excel(aa: list, site):
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    data_path_splice = os.path.join(origin_path, today)
    if not os.path.exists(data_path_splice):
        os.makedirs(data_path_splice)
    excel_data = os.path.join(data_path_splice, r'data.xlsx')
    if not os.path.exists(excel_data):
        wb = Workbook()
        wb.save(excel_data)
        wb.close()
    wb = openpyxl.load_workbook(excel_data)
    sheet = wb.active
    site_dict = {'EU': [1, 2], 'US': [3, 4], 'JP': [5, 6]}
    now_site = site_dict[site]
    for index, data01 in enumerate(aa):
        data = str(data01)
        sheet.cell(row=index+1, column=now_site[0]).value = data
        re_space = re.sub(r'[" "]', '', data)
        sheet.cell(row=index + 1, column=now_site[1]).value = re_space
    wb.save(excel_data)


write_to_excel([11,33,455], 'JP')